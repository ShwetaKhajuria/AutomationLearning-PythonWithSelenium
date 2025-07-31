import openpyxl


#Path to consider :- File--> Workbook--> Sheet1--> rows --> Cells
#File
File = 'C:/Users/khajuria_S/Desktop/Test Data for Selenium/Test_Data.xlsx'
#Workbook
workbook=openpyxl.load_workbook(File)
#Sheet1
sheet=workbook['Sheet1']
#rows
rows =sheet.max_row #calculate Number of rows
#Columns
columns = sheet.max_column #calculate Number of columns

#reading all rows
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='          ')
    print()


