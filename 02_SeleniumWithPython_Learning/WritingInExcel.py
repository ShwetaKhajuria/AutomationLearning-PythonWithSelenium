import openpyxl
from openpyxl.workbook import Workbook



#Path to consider :- File--> Workbook--> Sheet1--> rows --> Cells
# File ="C:/Users/khajuria_S/Desktop/Test Data for Selenium/Empty_Excel.xlsx"
# workbook = openpyxl.load_workbook(File)
# sheet = workbook['Sheet1']
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
#         workbook.save(File)

#Diffrent data
#Path to consider :- File--> Workbook--> Sheet1--> rows --> Cells
File = "C:/Users/khajuria_S/Desktop/Test Data for Selenium/Empty_Excel.xlsx"
workbook = openpyxl.load_workbook(File)
sheet = workbook['Sheet1']

sheet.cell(1,1).value = "Book Name"
sheet.cell(1,2).value="Price"

sheet.cell(2,1).value = "Java"
sheet.cell(2,2).value= "$50"

sheet.cell(3,1).value="Python"
sheet.cell(3,2).value = "$30"

workbook.save(File)